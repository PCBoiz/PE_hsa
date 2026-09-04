"""Đọc một bảng tính thành các DÒNG — nơi duy nhất trong repo ĐỌC định dạng ấy.

Nói "đọc" chứ không nói "biết", vì hai chỗ khác cũng biết: `mockexam/quan_tri.py`
DỰNG mẫu `.xlsx` và `teaching/exports.py` GHI `.csv`. Câu cũ ở đây nhận độc
quyền cho cả ba việc — một lời hứa rộng hơn thứ tệp này thật sự giữ.

── VÌ SAO CÓ TỆP NÀY (04/09/2026) ──────────────────────────────────────────

Anh chốt đề thi và giáo trình do TopHSA cung cấp. Đường nhập hiện có nhận
**JSON** (`docs/NHAP_GIAO_TRINH.md`) — định dạng của lập trình viên. Người soạn
đề ở một trung tâm luyện thi làm việc trong Excel, không trong dấu ngoặc nhọn;
bảo họ chuyển sang JSON là dựng một bức tường ngay ở chỗ nội dung phải chảy vào.

Cách làm lấy từ Kahoot (tra 04/09/2026): **tải mẫu → điền → tải lên → sai thì
báo TỪNG DÒNG kèm số dòng để sửa rồi nạp lại.** Ba chi tiết của họ đáng chép:

  · trần độ dài hiện ra dưới dạng LỖI chứ không cắt ngầm — cắt ngầm thì người
    soạn tưởng đã vào đủ, và phát hiện ra vào lúc học viên đang thi;
  · lỗi nêu đủ MỌI dòng sai trong một lượt, không dừng ở dòng đầu tiên — sửa
    một lỗi rồi tải lên lại để gặp lỗi thứ hai là một vòng lặp làm người ta bỏ
    cuộc;
  · không có trạng thái nửa vời: kiểm hết rồi mới ghi.

Hai điều sau repo này đã có sẵn ở đường JSON (`lessons/content.py`,
`courseadmin/views.py`), nên tệp này chỉ lo phần ĐỌC.

── VÌ SAO ĐỌC CẢ .xlsx LẪN .csv ────────────────────────────────────────────

`.xlsx` là thứ người ta thật sự có, và nó không có vấn đề mã hoá: ô chứa chuỗi
Unicode, không phải byte. `.csv` thì Excel bản tiếng Việt lưu mặc định theo bảng
mã hệ thống, nên "Định lượng" quay về thành ký tự hỏng — repo này đã trả giá cho
đúng chuyện đó ở đường XUẤT (phải thêm BOM). Nhận cả hai, nhưng mẫu tải về là
`.xlsx` để người dùng đi đường ít bẫy nhất.
"""
import csv
import datetime
import io
import re

#: Trần số dòng một lần nhập. Không phải để tiết kiệm bộ nhớ — để một file dán
#: nhầm (cả cuốn ngân hàng câu hỏi) không âm thầm thành 5.000 câu hỏi.
MAX_DONG = 2000

#: Trần độ dài một ô. Vượt thì BÁO, không cắt. Xem lý do ở đầu tệp.
MAX_O = 4000


class LoiBangTinh(Exception):
    """Không đọc nổi tệp — khác hẳn với 'đọc được nhưng nội dung sai'."""


def _chuan_tieu_de(s):
    """Tên cột về dạng so khớp được: thường, bỏ khoảng trắng thừa.

    KHÔNG bỏ dấu tiếng Việt: tiêu đề mẫu là tiếng Việt và người dùng sẽ copy
    nguyên, nên so khớp thẳng là đúng và dễ đoán. Bỏ dấu ở đây sẽ làm "Đáp án"
    và "Dap an" cùng khớp — nghe tiện, nhưng rồi một cột gõ sai chính tả cũng
    khớp và người soạn không bao giờ biết cột nào đã được dùng.
    """
    return ' '.join(str(s or '').split()).lower()


class _LoiO(Exception):
    """Một Ô không đọc lại được trung thành. `doc_xlsx` bắt và gắn thêm toạ độ.

    Có ngoại lệ riêng vì `_o` biết CÁI GÌ sai còn `doc_xlsx` mới biết Ô NÀO.
    Trả về một chuỗi báo lỗi thì nó sẽ trôi vào dữ liệu như một đáp án.
    """


def _sach_dinh_dang(fmt):
    """Bỏ phần NGUYÊN VĂN khỏi chuỗi định dạng, chỉ còn các MÃ định dạng.

    Excel cho phép hiện một ký tự nguyên văn bằng hai cách: thoát bằng `\\` hoặc
    bọc trong nháy kép. Cả hai đều KHÔNG mang nghĩa định dạng.
    """
    con = re.sub(r'"[^"]*"', '', fmt or '')
    return re.sub(r'\\.', '', con)


def _la_phan_tram(fmt):
    """Định dạng PHẦN TRĂM (nhân 100) — không phải "có ký tự % ở đâu đó".

    Đo 04/09/2026, bản đầu chỉ hỏi `'%' in fmt` và sai ở hai định dạng RẤT hay
    dùng — chính là thủ thuật để hiện dấu `%` mà không muốn Excel chia 100:

        ô 30, định dạng `0\\%`    → Excel hiện "30%"  → bản đầu đọc ra "3000%"
        ô 30, định dạng `0" %"`  → Excel hiện "30 %" → bản đầu đọc ra "3000%"

    Tức bản vá phần trăm sáng nay tự tạo ra một đường sai gấp 100 lần ở đúng
    những ô mà người soạn đã cẩn thận nhất.
    """
    return '%' in _sach_dinh_dang(fmt)


def _la_phan_so(fmt):
    """Định dạng PHÂN SỐ (`# ?/?`, `?/?`, `# ??/??`).

    Phải soi mã định dạng chứ không tìm dấu `/` trần: `dd/mm/yyyy` cũng có `/`.
    Mẫu `[?#0] / [?#0]` chỉ khớp phân số vì `d`, `m`, `y` không phải mã số.
    """
    return bool(re.search(r'[?#0]\s*/\s*[?#0]', _sach_dinh_dang(fmt)))


def _o(v, dinh_dang=None):
    """Một ô về chuỗi — CÁI NGƯỜI SOẠN NHÌN THẤY, không phải cái Excel lưu.

    Hai chỗ hai thứ ấy khác nhau, và cả hai hỏng ÂM THẦM (đo 04/09/2026):

        ô hiện "30%"        → Excel lưu 0.3        → bản cũ trả "0.3"
        ô hiện "04/09/2026" → Excel lưu datetime   → "2026-09-04 00:00:00"

    Không chỗ nào báo lỗi: chuỗi ấy hợp lệ, nó chỉ SAI. Người soạn gõ "30%" vào
    ô đáp án, và học viên phải trả lời "0.3" mới được tính đúng — lộ ra vào đúng
    lúc đang thi. Đây là loại lỗi mà bộ kiểm nội dung không thể bắt được, vì tới
    lúc nó nhìn thì dữ liệu đã sai từ trước rồi.

    `number_format` phân biệt được hai thứ, và nó CÓ ở chế độ `read_only` — đã
    đo — nhưng chỉ khi đọc bằng `values_only=False`, tức phải bỏ đường tắt cũ.
    """
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        # Excel không có kiểu "chỉ ngày": ngày về là datetime lúc 00:00.
        if v.hour or v.minute or v.second:
            return v.strftime('%d/%m/%Y %H:%M')
        return v.strftime('%d/%m/%Y')
    if isinstance(v, datetime.date):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if _la_phan_so(dinh_dang):
            # TỪ CHỐI, không đoán. Excel lưu `1/2` là số thực 0.5; dựng lại
            # chuỗi người soạn NHÌN THẤY thì được `1/2`, nhưng `2 1/2` (phân số
            # hỗn) lại ra `5/2` — cùng giá trị, KHÁC CHUỖI, và máy chấm so
            # chuỗi. Đọc lại gần đúng một đáp án còn tệ hơn từ chối nó: từ chối
            # thì người soạn sửa trong 10 giây, đọc sai thì cả lớp mất điểm.
            raise _LoiO('ô đang định dạng Phân số. Hệ thống không đọc lại được '
                        'đúng thứ bạn nhìn thấy — định dạng cột là Văn bản (Text) '
                        'rồi gõ lại, ví dụ 1/2.')
        if _la_phan_tram(dinh_dang):
            pt = v * 100
            if abs(pt - round(pt)) < 1e-9:
                return '%d%%' % round(pt)
            return '%s%%' % ('%.4f' % pt).rstrip('0').rstrip('.')
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
    # KHOẢNG TRẮNG KHÔNG NGẮT (` `) → dấu cách thường.
    #
    # `.strip()` cắt được nó ở hai đầu nhưng không đụng tới nó ở GIỮA. Dán từ
    # một trang web vào Excel thì `&nbsp;` đi theo, và ô hiện "Hà Nội" y hệt ô
    # bình thường — mắt không phân biệt được.
    #
    # Máy chấm (`mockexam/views.py`) chuẩn hoá bằng `.replace(' ', '')`, tức chỉ
    # bỏ dấu cách ASCII. Một đáp án ĐIỀN mang ` ` ở giữa là đáp án KHÔNG AI
    # trả lời đúng được — 100% học viên sai câu đó, và không chỗ nào báo.
    return str(v).replace(' ', ' ').strip()


def doc_xlsx(du_lieu):
    import openpyxl  # nhập trễ: chỉ đường nhập mới cần, không nặng lúc khởi động

    try:
        wb = openpyxl.load_workbook(io.BytesIO(du_lieu), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — mọi lỗi đọc tệp đều là "tệp hỏng"
        raise LoiBangTinh('Không mở được tệp .xlsx: %s' % e) from e

    def doc_trang(ws):
        """Một trang → danh sách dòng. Ô nào không đọc lại được thì BÁO KÈM TOẠ ĐỘ.

        Bọc cả vòng lặp trong `try`: ở chế độ `read_only`, XML của trang được
        phân tích LƯỜI, tức lỗi tệp hỏng nổ ra ở ĐÂY chứ không ở `load_workbook`.
        Bản đầu chỉ ôm `load_workbook`, nên một tệp tải về dở dang (Wi-Fi rớt)
        cho `ParseError` lọt ra ngoài `except LoiBangTinh` của view → HTTP 500,
        tức "lỗi máy chủ" cho một tệp hỏng của người dùng.
        """
        ra = []
        try:
            for hang in ws.iter_rows():
                dong = []
                for c in hang:
                    try:
                        dong.append(_o(c.value, getattr(c, 'number_format', None)))
                    except _LoiO as e:
                        raise LoiBangTinh('Ô %s (trang "%s"): %s'
                                          % (c.coordinate, ws.title, e)) from e
                ra.append(dong)
        except LoiBangTinh:
            raise
        except (NameError, AttributeError, ImportError):
            # ĐỪNG hoá trang lỗi lập trình thành "tệp của bạn hỏng". Đã xảy ra
            # ngay trong phiên viết ra khối này: thiếu `import re` biến thành
            # thông báo "Đọc tới giữa chừng thì tệp .xlsx hỏng" và suýt nữa tôi
            # đi sửa tệp mẫu. Ba loại này không bao giờ do dữ liệu người dùng.
            raise
        except Exception as e:  # noqa: BLE001 — XML hỏng, ô lỗi, kiểu lạ…
            raise LoiBangTinh('Đọc tới giữa chừng thì tệp .xlsx hỏng (%s: %s). '
                              'Mở lại bằng Excel, lưu lại, rồi tải lên.'
                              % (type(e).__name__, e)) from e
        return ra

    # Lấy trang ĐẦU TIÊN CÓ DỮ LIỆU, không cứng `worksheets[0]`. Người ta hay để
    # trang 1 là "Hướng dẫn" rồi mới tới dữ liệu ở trang 2; bản cũ đọc trang 1
    # rồi báo "thiếu cột bắt buộc" trong khi dữ liệu nằm ngay trang bên cạnh —
    # một thông báo lỗi chỉ đúng chữ, còn chỗ nó chỉ tới thì không có gì sai.
    #
    # Cần ÍT NHẤT 2 dòng (tiêu đề + một dòng dữ liệu): một trang chỉ có một dòng
    # chữ giới thiệu không phải là bảng dữ liệu.
    for i, ws in enumerate(wb.worksheets):
        rows = doc_trang(ws)
        if sum(1 for h in rows if any(o for o in h)) >= 2:
            _chan_cong_thuc_rong(du_lieu, i, rows)
            return rows
    return doc_trang(wb.worksheets[0]) if wb.worksheets else []


def _chan_cong_thuc_rong(du_lieu, chi_so_trang, rows):
    """Ô CÔNG THỨC chưa được Excel tính sẵn → báo, đừng để nó thành ô trống.

    `data_only=True` trả giá trị đã lưu sẵn trong tệp. Excel lưu sẵn giá trị ấy,
    nhưng tệp do thư viện khác sinh ra (hoặc do LibreOffice/Google Sheets xuất)
    thì có khi không — và lúc ấy openpyxl trả ``None``, **không phân biệt được
    với một ô trống thật** (đã đo: cả hai đều `value=None, data_type='n'`).

    Hậu quả trước bản vá: một cột "Lựa chọn D" viết bằng công thức biến mất im
    lặng, câu vẫn vào ngân hàng đề với 3 phương án. Và nếu cột Đáp án là công
    thức thì câu bị báo "thiếu đáp án" — một thông báo chỉ đường sai.

    Phải nạp LẠI tệp với ``data_only=False`` mới thấy công thức. Chỉ làm khi
    thật sự có ô rỗng: tệp không có ô rỗng nào thì không có gì để nhầm.
    """
    if not any('' in h for h in rows):
        return
    import openpyxl

    try:
        wb2 = openpyxl.load_workbook(io.BytesIO(du_lieu), read_only=True, data_only=False)
        ws2 = wb2.worksheets[chi_so_trang]
        thu_pham = []
        for hang in ws2.iter_rows():
            for c in hang:
                if isinstance(c.value, str) and c.value.startswith('='):
                    r, cot = c.row - 1, c.column - 1
                    if r < len(rows) and cot < len(rows[r]) and rows[r][cot] == '':
                        thu_pham.append(c.coordinate)
    except Exception:  # noqa: BLE001 — lượt nạp phụ hỏng thì đừng làm hỏng lượt chính
        return
    if thu_pham:
        raise LoiBangTinh(
            'Ô %s là CÔNG THỨC nhưng tệp chưa lưu kết quả của nó, nên hệ thống '
            'đọc ra ô trống. Mở tệp bằng Excel, nhấn Ctrl+S để Excel tính và lưu '
            'lại kết quả — hoặc chép rồi "Dán giá trị" (Paste Values).'
            % ', '.join(thu_pham[:8]))


def doc_csv(du_lieu):
    for ma in ('utf-8-sig', 'utf-8', 'cp1258', 'latin-1'):
        try:
            van = du_lieu.decode(ma)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise LoiBangTinh('Không đọc được mã ký tự của tệp .csv. '
                          'Lưu lại bằng "CSV UTF-8" rồi thử lại, hoặc dùng .xlsx.')
    # `Sniffer` đoán dấu phân cách: Excel bản Việt hay dùng dấu chấm phẩy.
    try:
        dau = csv.Sniffer().sniff(van[:4000], delimiters=',;\t').delimiter
    except csv.Error:
        dau = ','
    try:
        return [[_o(c) for c in hang] for hang in csv.reader(io.StringIO(van), delimiter=dau)]
    except csv.Error as e:
        # `csv.reader` có trần ô cứng 131.072 ký tự và ném `_csv.Error` —
        # KHÔNG phải `LoiBangTinh`, nên nó lọt ra ngoài `except LoiBangTinh` của
        # view và thành HTTP 500. Dán một đoạn văn dài vào ô "Giải thích" là đủ:
        # người dùng nhận "lỗi máy chủ" cho một việc họ sửa được trong 5 giây.
        #
        # KHÔNG nới `csv.field_size_limit`: trần của repo là `MAX_O` (4.000),
        # chặt hơn nhiều, nên nới lên chỉ đổi chỗ báo lỗi chứ không đổi kết quả.
        raise LoiBangTinh('Tệp .csv có một ô quá dài (trần %d ký tự mỗi ô). '
                          'Rút ngắn ô đó rồi tải lại. (%s)' % (MAX_O, e)) from e


def doc(ten_tep, du_lieu):
    """Đọc tệp → danh sách các DÒNG (mỗi dòng là danh sách ô, đã về chuỗi)."""
    ten = (ten_tep or '').lower()
    if ten.endswith('.xlsx'):
        return doc_xlsx(du_lieu)
    if ten.endswith(('.csv', '.txt')):
        return doc_csv(du_lieu)
    raise LoiBangTinh('Chỉ nhận tệp .xlsx hoặc .csv (đang nhận: %s).'
                      % (ten_tep or 'không rõ tên'))


def thanh_ban_ghi(hang, cot_bat_buoc=(), ten_khac=None):
    """Dòng thô → danh sách ``(so_dong, {tên cột: giá trị})``.

    ``so_dong`` là số dòng NHƯ NGƯỜI DÙNG THẤY TRONG EXCEL (tính từ 1, kể cả
    dòng tiêu đề). Đây là chi tiết quyết định việc sửa lỗi có nhanh không: báo
    "dòng thứ 3 của mảng" thì người soạn phải tự đếm; báo "dòng 4" thì họ nhấn
    Ctrl+G trong Excel là tới nơi.

    ``ten_khac`` cho phép một cột có nhiều tên gọi (ví dụ "đáp án" và "dap an"),
    khai TƯỜNG MINH chứ không đoán bằng cách bỏ dấu.
    """
    # ĐÁNH SỐ TRƯỚC, LỌC SAU. Bản đầu làm ngược lại — lọc dòng trống rồi mới
    # `enumerate` — nên mỗi dòng trống trong tệp đẩy toàn bộ số dòng phía dưới
    # lệch đi một. Docstring ngay trên đây HỨA "số dòng như người dùng thấy
    # trong Excel"; mã thì đếm theo vị trí trong mảng đã lọc. Chú thích sai
    # nguy hiểm hơn không có chú thích: nó tắt phản xạ kiểm tra của người sau,
    # và ở đây nó còn gửi người soạn tới sửa NHẦM DÒNG.
    danh = [(i, h) for i, h in enumerate(hang, start=1) if any(o for o in h)]
    if not danh:
        raise LoiBangTinh('Tệp không có dòng nào.')

    tieu_de = [_chuan_tieu_de(o) for o in danh[0][1]]
    if ten_khac:
        tieu_de = [ten_khac.get(t, t) for t in tieu_de]

    # HAI CỘT CÙNG TÊN → cột sau ghi đè cột trước, IM LẶNG (vá 04/09/2026).
    #
    # Nó không cần hai tiêu đề giống hệt nhau: `ten_khac` gộp "đáp án đúng" về
    # "đáp án", và một cột phụ tên vỏn vẹn "A" về "lựa chọn a". Nên bảng
    #
    #     Phần thi | Câu hỏi | Lựa chọn A | Lựa chọn B | Đáp án | Đáp án đúng
    #     Định lượng | 2+2? |     3      |     4      |   4    |      3
    #
    # (cột cuối là ghi chú thừa của người soạn) ghi vào ngân hàng đề `answer='3'`
    # — SAI, không một chữ báo. Cả lớp chọn 4 sẽ bị trừ điểm.
    #
    # Kiểu thứ hai: copy-paste nhân đôi một cột "Lựa chọn C" rồi để trống bản
    # sao — phương án C biến mất khỏi câu hỏi.
    #
    # Không đoán cột nào là cột thật: hỏi người soạn. Đây là đúng chỗ mà một câu
    # hỏi rẻ hơn một phỏng đoán rất nhiều.
    dem = {}
    for t in tieu_de:
        if t:
            dem[t] = dem.get(t, 0) + 1
    trung = sorted(t for t, n in dem.items() if n > 1)
    if trung:
        raise LoiBangTinh(
            'Có hai cột cùng tên: %s. Xoá bớt cột thừa rồi tải lại — hệ thống '
            'KHÔNG tự đoán cột nào là cột thật, vì đoán sai thì đáp án sai mà '
            'không ai biết. Lưu ý vài tên khác nhau tính là MỘT cột (ví dụ '
            '"Đáp án" và "Đáp án đúng", hoặc "A" và "Lựa chọn A").'
            % ', '.join('"%s"' % t for t in trung))

    thieu = [c for c in cot_bat_buoc if c not in tieu_de]
    if thieu:
        raise LoiBangTinh(
            'Thiếu cột bắt buộc: %s. Dòng đầu của tệp phải là dòng TIÊU ĐỀ, '
            'chép nguyên từ mẫu tải về. Đang thấy: %s.'
            % (', '.join('"%s"' % c for c in thieu),
               ', '.join('"%s"' % t for t in tieu_de if t) or '(không có tiêu đề nào)'))

    than = danh[1:]
    if len(than) > MAX_DONG:
        raise LoiBangTinh('Tối đa %d dòng một lần nhập (tệp đang có %d). '
                          'Chia nhỏ ra rồi nhập nhiều lượt.' % (MAX_DONG, len(than)))

    ra = []
    for so_dong, h in than:
        ban = {}
        for j, ten in enumerate(tieu_de):
            if ten:
                o = h[j] if j < len(h) else ''
                if len(o) > MAX_O:
                    raise LoiBangTinh(
                        'Dòng %d, cột "%s": ô dài %d ký tự, tối đa %d. '
                        'Cắt bớt trong Excel rồi nhập lại — hệ thống KHÔNG tự cắt, '
                        'vì cắt ngầm thì nội dung mất mà không ai biết.'
                        % (so_dong, ten, len(o), MAX_O))
                ban[ten] = o
        ra.append((so_dong, ban))
    return ra

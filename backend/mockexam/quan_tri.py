"""Quản trị đề thi thử — liệt kê, nhập từ bảng tính, tải mẫu.

Tới 04/09/2026 KHÔNG có API quản trị nào cho đề thi: đề duy nhất trên CSDL đến
từ `seed_data`, và đường duy nhất để có đề thứ hai là sửa mã nguồn. Xem
`mockexam/nhap.py` về việc vì sao chuyện đó chặn đúng trụ cột 4 của sản phẩm.

QUYỀN: `IsContentEditor` — quản trị viên hoặc `Biên tập nội dung`. Cùng ranh
giới với soạn giáo trình: đề thi là NỘI DUNG, không phải dữ liệu học viên.
"""
import io
import json

from django.db import transaction
from django.http import HttpResponse
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from common import audit
from common.bangtinh import LoiBangTinh, doc, thanh_ban_ghi
from common.db import q, q1, x
from common.permissions import IsContentEditor
from mockexam.nhap import COT_BAT_BUOC, TEN_KHAC, TIEU_DE_MAU, doc_cau_hoi

#: Trần kích thước tệp. Một bảng tính câu hỏi thật nặng vài trăm kB; 5 MB là
#: rộng rãi mà vẫn chặn được việc dán nhầm một tệp ảnh vào ô tải lên.
MAX_BYTES = 5 * 1024 * 1024

#: Dòng mẫu, để người soạn thấy NGAY hình dạng thay vì đọc tài liệu. Kahoot làm
#: đúng thế và đó là chi tiết khiến mẫu của họ dùng được mà không cần hướng dẫn.
DONG_MAU = [
    ['Định lượng', 'Giá trị của biểu thức 2³ + 3² bằng bao nhiêu?',
     '15', '16', '17', '18', '17', 'ql_01', 'Số học', '2³=8, 3²=9, tổng là 17.'],
    ['Định lượng', 'Một số tăng 20% rồi giảm 20% thì so với ban đầu?',
     'Không đổi', 'Tăng 4%', 'Giảm 4%', 'Giảm 20%', 'C', 'ql_02', 'Tỉ lệ',
     'Nhân hệ số: 1,2 × 0,8 = 0,96 → giảm 4%.'],
    ['Định tính', 'Điền số còn thiếu: 2, 4, 8, 16, …', '', '', '', '',
     '32', '', 'Dãy số', 'Bỏ trống hết cột Lựa chọn thì thành câu ĐIỀN đáp án.'],
]

# `TIEU_DE_MAU` nay NHẬP TỪ `mockexam/nhap.py` — xem `nhap.COT`. Trước 04/09
# (chiều) nó là một mảng gõ tay ở đây, ngay dưới một chú thích nói rằng nó "lấy
# thẳng từ hằng số mà bộ đọc dùng". Hai bản cho một danh sách, và câu chú thích
# canh cho chúng khớp thì chính nó là câu sai.


class _Base(APIView):
    permission_classes = [IsContentEditor]


class AdminMockExamsView(_Base):
    """GET /api/admin/mock-exams — danh sách đề, kèm số lượt đã làm."""

    def get(self, request):
        rows = q('''SELECT e.id, e.title, e.duration_minutes, e.total_questions,
                           e.is_published, e.created_at,
                           (SELECT COUNT(*) FROM mock_attempts a WHERE a.exam_id = e.id) AS luot
                    FROM mock_exams e ORDER BY e.id DESC''')
        return Response({'exams': [{
            'id': r['id'], 'title': r['title'],
            'durationMinutes': r['duration_minutes'],
            'totalQuestions': r['total_questions'],
            'isPublished': r['is_published'],
            'attempts': r['luot'],
        } for r in rows]})


class AdminMockExamTemplateView(_Base):
    """GET /api/admin/mock-exams/template.xlsx — mẫu để điền.

    Sinh tại chỗ chứ không để một tệp tĩnh trong repo: mẫu và bộ đọc phải luôn
    khớp nhau, và một tệp tĩnh là bản thứ hai sẽ trôi. Tên cột ở đây lấy thẳng
    từ hằng số mà bộ đọc dùng.
    """

    def get(self, request):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Câu hỏi'
        ws.append(TIEU_DE_MAU)
        for d in DONG_MAU:
            ws.append(d)
        for i, w in enumerate([14, 60, 18, 18, 18, 18, 16, 12, 16, 46], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        huong = wb.create_sheet('Hướng dẫn')
        for dong in [
            ['Mỗi DÒNG là một câu hỏi. Giữ nguyên dòng tiêu đề ở trang "Câu hỏi".'],
            [''],
            ['Phần thi', 'Định lượng · Định tính · Khoa học (bắt buộc)'],
            ['Câu hỏi', 'Nội dung câu hỏi (bắt buộc)'],
            ['Lựa chọn A–D', 'Bỏ trống HẾT bốn cột này thì câu thành dạng ĐIỀN đáp án'],
            ['Đáp án', 'Chép nguyên văn phương án đúng, HOẶC ghi chữ cái A/B/C/D'],
            ['Mã câu', 'Không bắt buộc. Bỏ trống thì hệ thống tự đặt.'],
            ['Chủ đề', 'Không bắt buộc. Dùng cho phân tích mạnh–yếu.'],
            ['Giải thích', 'Không bắt buộc. Được lưu lại cho lần engine hiện lời giải.'],
            [''],
            ['Sai ở đâu, hệ thống báo ĐÚNG SỐ DÒNG như trong Excel.'],
            ['Chỉ cần một dòng sai là KHÔNG dòng nào được ghi — không có trạng thái nửa vời.'],
        ]:
            huong.append(dong)
        huong.column_dimensions['A'].width = 16
        huong.column_dimensions['B'].width = 76

        buf = io.BytesIO()
        wb.save(buf)
        res = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = 'attachment; filename="mau-de-thi-thu.xlsx"'
        return res


class AdminMockExamImportView(_Base):
    """POST /api/admin/mock-exams/import — nhập một đề từ bảng tính.

    Nhận `multipart/form-data`: `file` + `title` + `duration_minutes`, và
    `exam_id` nếu muốn GHI ĐÈ một đề đã có thay vì tạo đề mới.

    KIỂM HẾT RỒI MỚI GHI. Một dòng sai là không dòng nào được ghi — cùng luật
    với đường nhập giáo trình, và vì cùng lý do: một đề nhập nửa chừng trông y
    hệt một đề soạn thiếu, và không ai biết nó thiếu ở đâu.
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        tep = request.FILES.get('file')
        if not tep:
            return Response({'error': 'Chưa chọn tệp bảng tính.'}, status=400)
        if tep.size > MAX_BYTES:
            return Response({'error': 'Tệp nặng %.1f MB, tối đa %d MB.'
                                      % (tep.size / 1048576, MAX_BYTES // 1048576)}, status=400)

        try:
            hang = doc(tep.name, tep.read())
            ban_ghi = thanh_ban_ghi(hang, cot_bat_buoc=COT_BAT_BUOC, ten_khac=TEN_KHAC)
        except LoiBangTinh as e:
            return Response({'error': str(e)}, status=400)

        cau, loi = doc_cau_hoi(ban_ghi)
        if loi:
            return Response({'error': '%d lỗi trong tệp — chưa ghi gì cả.' % len(loi),
                             'details': loi[:50]}, status=400)

        title = (request.data.get('title') or '').strip()
        if not title:
            return Response({'error': 'Đề thi phải có tên.'}, status=400)
        try:
            phut = int(request.data.get('duration_minutes') or 0)
        except (TypeError, ValueError):
            phut = 0
        if not 1 <= phut <= 600:
            return Response({'error': 'Thời gian làm bài phải là số phút trong khoảng 1–600.'},
                            status=400)

        # `mock_exams.id` là INTEGER, còn `request.data` của một biểu mẫu nhiều
        # phần thì toàn chuỗi. Chuyền thẳng vào truy vấn thì `exam_id='abc'` làm
        # psycopg ném `invalid input syntax for type integer` — tức 500, tức
        # trang báo "lỗi máy chủ" cho một việc người dùng gõ sai. 400 kèm câu
        # giải thích mới là đúng chuyện đang xảy ra.
        tho_exam_id = request.data.get('exam_id') or None
        exam_id = None
        if tho_exam_id is not None:
            try:
                exam_id = int(tho_exam_id)
            except (TypeError, ValueError):
                return Response({'error': 'exam_id phải là số (mã đề cần ghi đè). '
                                          'Bỏ trống nếu muốn tạo đề mới.'}, status=400)

        than = json.dumps(cau, ensure_ascii=False)

        with transaction.atomic():
            if exam_id:
                cu = q1('SELECT id, title, is_published FROM mock_exams WHERE id=%s',
                        (exam_id,))
                if not cu:
                    return Response({'error': 'Không tìm thấy đề thi này.'}, status=404)
                # Đề đã có người làm thì ĐỔI ĐỀ là làm hỏng ý nghĩa của các lượt
                # đã chấm: điểm cũ tính trên bộ câu hỏi khác. Chặn thẳng và nói
                # rõ đường đi tiếp, thay vì âm thầm ghi đè.
                luot = q1('SELECT COUNT(*) AS n FROM mock_attempts WHERE exam_id=%s',
                          (exam_id,))['n']
                if luot:
                    return Response(
                        {'error': 'Đề này đã có %d lượt làm bài. Ghi đè sẽ làm điểm của '
                                  'các lượt ấy không còn ý nghĩa (chúng chấm trên bộ câu '
                                  'hỏi cũ). Tạo đề MỚI, rồi ẩn đề cũ đi.' % luot},
                        status=409)
                x('''UPDATE mock_exams SET title=%s, duration_minutes=%s,
                         total_questions=%s, questions_json=%s::jsonb WHERE id=%s''',
                  (title, phut, len(cau), than, exam_id))
                hanh_dong, ma = audit.MOCK_EXAM_UPDATE, exam_id
                # Ghi đè KHÔNG đụng tới `is_published` — nên phải đọc giá trị
                # thật, không suy ra. Bản đầu trả `published: bool(exam_id)`,
                # tức "đã ghi đè" bị trình bày thành "đã xuất bản": một đề đang
                # ẩn vẫn báo về `published: true`, và người soạn tin là học viên
                # đã thấy đề rồi trong khi chưa ai thấy gì.
                da_hien = bool(cu['is_published'])
            else:
                row = q1('''INSERT INTO mock_exams
                                (title, description, duration_minutes, total_questions,
                                 questions_json, is_published)
                            VALUES (%s, '', %s, %s, %s::jsonb, FALSE) RETURNING id''',
                         (title, phut, len(cau), than))
                hanh_dong, ma = audit.MOCK_EXAM_CREATE, row['id']
                da_hien = False        # đề mới luôn ẩn cho tới khi bấm xuất bản

            audit.record(request, hanh_dong, target_type='mock_exam', target_id=ma,
                         target_label=title,
                         summary='%s đề thi thử "%s" — %d câu, %d phút.'
                                 % ('Cập nhật' if exam_id else 'Nhập', title, len(cau), phut),
                         detail={'questions': len(cau), 'minutes': phut,
                                 'file': tep.name})

        return Response({'ok': True, 'id': ma, 'questions': len(cau),
                         'published': da_hien}, status=200 if exam_id else 201)


class AdminMockExamPublishView(_Base):
    """PATCH /api/admin/mock-exams/<id>/publish — bật/tắt hiển thị cho học viên.

    Đề nhập vào mặc định CHƯA xuất bản: người soạn phải mở ra xem trước rồi mới
    bật. Nhập xong là học viên thấy ngay thì một lần nhập nhầm cột đáp án sẽ
    chấm sai cho cả lớp trước khi ai kịp nhận ra.
    """

    def patch(self, request, exam_id):
        cu = q1('SELECT id, title, is_published FROM mock_exams WHERE id=%s', (exam_id,))
        if not cu:
            return Response({'error': 'Không tìm thấy đề thi này.'}, status=404)
        bat = bool(request.data.get('published'))
        x('UPDATE mock_exams SET is_published=%s WHERE id=%s', (bat, exam_id))
        audit.record(request, audit.MOCK_EXAM_PUBLISH, target_type='mock_exam',
                     target_id=exam_id, target_label=cu['title'],
                     summary='%s đề thi thử "%s".'
                             % ('Xuất bản' if bat else 'Ẩn', cu['title']),
                     detail={'published': bat})
        return Response({'ok': True, 'published': bat})

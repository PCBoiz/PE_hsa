"""Phép kiểm cho phòng thi thử — đồng hồ ở máy chủ, một lượt vào sổ (31/08/2026).

Khu này trước nay không có tệp test nào. Đo được trên bản chạy thật: nộp RỖNG →
nhận đủ 9 đáp án → nộp lại → 9/9, +100 XP; `duration_seconds` nhận cả −5000;
`started_at` có cột nhưng không dòng mã nào ghi.

Mọi phép kiểm ở đây chạy trong giao dịch được cuộn lại (conftest.py), nên không
để lại dòng nào trên Neon — RULES §18.

`mockexam.views` nhập MUỘN trong từng phép kiểm: các phép kiểm quan trọng nhất
phải CHẠY ĐƯỢC trên mã CŨ (nơi `MockStartView` chưa tồn tại) để chứng minh chúng
đỏ ở đó vì đúng lý do, chứ không chết bằng ImportError.
"""
import json

import pytest
from rest_framework.test import APIRequestFactory, force_authenticate

from accounts.models import User
from common.db import q1, x

f = APIRequestFactory()

#: Đề tối giản: 2 câu định lượng, 1 câu định tính. Tự dựng chứ không mượn đề
#: thật — đề thật đổi nội dung thì phép kiểm phải vẫn đúng.
CAU_HOI = [
    {'id': 'q1', 'section': 'quantitative', 'type': 'mcq',
     'question': '2 + 2 = ?', 'options': ['3', '4', '5'], 'answer': '4'},
    {'id': 'q2', 'section': 'quantitative', 'type': 'fill',
     'question': '10% của 200 = ?', 'answer': '20'},
    {'id': 'q3', 'section': 'verbal', 'type': 'mcq',
     'question': 'Từ trái nghĩa với "cao"?', 'options': ['thấp', 'rộng'], 'answer': 'thấp'},
]
DUNG_HET = {'q1': '4', 'q2': '20', 'q3': 'thấp'}


# `.get(...)` chứ không `[...]` với các trường MỚI: trên mã cũ chúng không tồn
# tại, và ta cần phép kiểm đỏ bằng AssertionError nói rõ giá trị sai, chứ không
# phải KeyError — một màu đỏ không phân biệt được "thiếu tính năng" với "test hỏng".
def _goi(view, method, body=None, ai=None, **kw):
    req = (getattr(f, method)('/x', body, format='json') if body is not None
           else getattr(f, method)('/x'))
    force_authenticate(req, user=ai)
    return view.as_view()(req, **kw)


@pytest.fixture
def em(db):
    r = q1("INSERT INTO users (name, email, password, streak) "
           "VALUES ('HV Thi','hv_thi_tmp@example.com','x',0) RETURNING id")
    return User.objects.get(id=r['id'])


@pytest.fixture
def de(db):
    r = q1("INSERT INTO mock_exams (title, description, duration_minutes, "
           "total_questions, questions_json, is_published) "
           "VALUES ('De kiem tam','x',20,3,%s::jsonb,TRUE) RETURNING id",
           (json.dumps(CAU_HOI, ensure_ascii=False),))
    return r['id']


@pytest.fixture
def de_an(db):
    """Đề CHƯA xuất bản — không ai được nộp vào đây."""
    r = q1("INSERT INTO mock_exams (title, description, duration_minutes, "
           "total_questions, questions_json, is_published) "
           "VALUES ('De chua xuat ban','x',20,3,%s::jsonb,FALSE) RETURNING id",
           (json.dumps(CAU_HOI, ensure_ascii=False),))
    return r['id']


def _bat_dau(em, de):
    from mockexam.views import MockStartView
    return _goi(MockStartView, 'post', {}, ai=em, exam_id=de)


def _nop(em, de, answers, **them):
    from mockexam.views import MockSubmitView
    body = {'answers': answers}
    body.update(them)
    return _goi(MockSubmitView, 'post', body, ai=em, exam_id=de)


def _luu(em, de, answers):
    from mockexam.views import MockSaveView
    return _goi(MockSaveView, 'post', {'answers': answers}, ai=em, exam_id=de)


def _lui_dong_ho(attempt_id, giay):
    """Kéo `started_at` lùi lại — cách duy nhất để kiểm chuyện hết giờ."""
    x("UPDATE mock_attempts SET started_at = started_at - %s::interval WHERE id=%s",
      ('%d seconds' % giay, attempt_id))


# ── Luật 2: đáp án chỉ lộ theo từng câu ĐÃ trả lời ───────────────────────────

@pytest.mark.django_db
def test_nop_RONG_thi_KHONG_nhan_duoc_dap_an_nao(em, de):
    """Cách khai thác đã đo được: nộp rỗng → nhận cả 9 đáp án → nộp lại 9/9.

    Đỏ trên mã cũ: bản cũ luôn gán `'answer': correct` cho MỌI câu.
    """
    _bat_dau(em, de)
    r = _nop(em, de, {})
    assert r.status_code == 200
    lo = [c for c in r.data['results'] if c['answer'] is not None]
    assert lo == [], 'nộp rỗng vẫn nhận được đáp án của %d câu' % len(lo)
    assert r.data['score'] == 0


@pytest.mark.django_db
def test_cau_da_tra_loi_thi_MOI_hien_dap_an(em, de):
    """Xem lại bài vẫn phải dùng được — chỉ với câu mình đã làm."""
    _bat_dau(em, de)
    r = _nop(em, de, {'q1': '4'})
    theo_id = {c['id']: c for c in r.data['results']}
    assert theo_id['q1']['answer'] == '4' and theo_id['q1']['correct'] is True
    assert theo_id['q2']['answer'] is None and theo_id['q2']['answered'] is False
    assert theo_id['q3']['answer'] is None


# ── Luật 3 + 4: lượt nào vào sổ ─────────────────────────────────────────────

@pytest.mark.django_db
def test_luot_thu_hai_KHONG_cong_XP_va_KHONG_ghi_su_kien(em, de):
    """Anh Sơn chốt 31/08/2026: "một lượt tính điểm, làm lại không cộng XP".

    Đỏ trên mã cũ: bản cũ cộng `_mock_xp` cho MỌI lượt, không đếm lượt nào.
    """
    _bat_dau(em, de)
    mot = _nop(em, de, DUNG_HET)
    assert mot.data['score'] == 3
    assert mot.data.get('counted') is True
    assert mot.data.get('xpGained') > 0

    _bat_dau(em, de)
    hai = _nop(em, de, DUNG_HET)
    assert hai.data['score'] == 3, 'vẫn phải được chấm để luyện'
    assert hai.data.get('counted') is False
    assert hai.data.get('notCountedReason') == 'da_lam_roi'
    assert hai.data.get('xpGained') == 0

    n = q1("SELECT COUNT(*) AS n FROM learning_events "
           "WHERE user_id=%s AND kind='mock'", (em.id,))['n']
    assert n == 1, 'lượt luyện không được vào sổ điểm (thấy %d dòng)' % n


@pytest.mark.django_db
def test_nop_ma_KHONG_qua_start_thi_khong_vao_so(em, de):
    """Nửa còn lại của luật 1. Bỏ qua `/start` là bỏ qua toàn bộ giới hạn giờ:
    lấy đề, ngồi ba tiếng tra đáp án, rồi nộp. Bài vẫn được chấm và vẫn lưu để
    em không mất công, nhưng KHÔNG cộng gì.

    Đỏ cả trên mã cũ LẪN bản đầu ngày 31/08 — cả hai đều tính điểm cho nhánh này.
    """
    r = _nop(em, de, DUNG_HET, duration_seconds=900)
    assert r.data['score'] == 3, 'vẫn chấm để em không mất công'
    assert r.data.get('counted') is False
    assert r.data.get('notCountedReason') == 'khong_qua_dong_ho'
    assert r.data.get('xpGained') == 0
    n = q1("SELECT COUNT(*) AS n FROM learning_events "
           "WHERE user_id=%s AND kind='mock'", (em.id,))['n']
    assert n == 0
    dong = q1("SELECT counted, duration_seconds FROM mock_attempts "
              "WHERE user_id=%s ORDER BY id DESC LIMIT 1", (em.id,))
    assert dong['counted'] is False
    assert dong['duration_seconds'] == 0, 'không nhận con số thời lượng của trình duyệt'


@pytest.mark.django_db
def test_bo_do_toi_can_gio_roi_bat_dau_lai_thi_KHONG_duoc_tinh_diem_nua(em, de):
    """Mở đề, đọc hết, đóng tab, hôm sau bấm "Bắt đầu" lại.

    Lượt tính điểm bị TIÊU ngay khi MỞ — đề đã lộ cho em rồi. Nếu quyết định
    `counted` lúc NỘP thì đây là một cách làm lại vô hạn mà vẫn được tính điểm.
    """
    a = _bat_dau(em, de)
    assert a.data.get('counts') is True
    _lui_dong_ho(a.data['attemptId'], 20 * 60 + 300)

    b = _bat_dau(em, de)
    assert b.data['attemptId'] != a.data['attemptId'], 'phải là một lượt mới'
    assert b.data.get('counts') is False, 'lượt tính điểm đã bị tiêu khi mở lượt đầu'
    r = _nop(em, de, DUNG_HET)
    assert r.data.get('counted') is False
    assert r.data.get('xpGained') == 0


@pytest.mark.django_db
def test_luot_bo_do_van_duoc_cham_tren_phan_da_luu(em, de):
    """Máy sập giữa chừng thì kết quả phải là kết quả của bài làm dở, không
    phải một lượt biến mất."""
    a = _bat_dau(em, de)
    _luu(em, de, {'q1': '4', 'q2': '20'})
    _lui_dong_ho(a.data['attemptId'], 20 * 60 + 300)
    _bat_dau(em, de)   # lần vào lại đóng lượt cũ

    dong = q1("SELECT score, total, submitted_at FROM mock_attempts WHERE id=%s",
              (a.data['attemptId'],))
    assert dong['submitted_at'] is not None, 'lượt cạn giờ phải được đóng lại'
    assert (dong['score'], dong['total']) == (2, 3), dong


@pytest.mark.django_db
def test_CSDL_tu_chan_hai_luot_tinh_diem_cho_cung_mot_de(em, de):
    """Ở mức cách ly `read committed`, một câu SELECT rồi INSERT không chặn
    được năm request song song — chỉ ràng buộc duy nhất mới chặn được.

    Kiểm thẳng ràng buộc: INSERT tay một dòng tính điểm thứ hai phải bị từ chối.
    """
    from django.db import IntegrityError, transaction
    _bat_dau(em, de)
    with pytest.raises(IntegrityError):
        with transaction.atomic():
            x("INSERT INTO mock_attempts (user_id, exam_id, started_at, submitted_at, counted) "
              "VALUES (%s,%s,now(),now(),TRUE)", (em.id, de))


@pytest.mark.django_db
def test_CSDL_tu_chan_hai_luot_DANG_MO_cho_cung_mot_de(em, de):
    from django.db import IntegrityError, transaction
    _bat_dau(em, de)
    with pytest.raises(IntegrityError):
        with transaction.atomic():
            x("INSERT INTO mock_attempts (user_id, exam_id, started_at, submitted_at, counted) "
              "VALUES (%s,%s,now(),NULL,FALSE)", (em.id, de))


@pytest.mark.django_db
def test_luot_luyen_van_duoc_luu_de_xem_lai(em, de):
    """Không tính điểm KHÁC với không được làm: dòng vẫn phải có trong lịch sử."""
    from mockexam.views import MockAttemptsView
    _bat_dau(em, de)
    _nop(em, de, {'q1': '4'})
    _bat_dau(em, de)
    _nop(em, de, {'q1': '4'})
    r = _goi(MockAttemptsView, 'get', ai=em)
    cua_de = [a for a in r.data['attempts'] if a['exam_id'] == de]
    assert len(cua_de) == 2
    assert [a.get('counted') for a in cua_de] == [False, True]


# ── Luật 1: đồng hồ thuộc về máy chủ ────────────────────────────────────────

@pytest.mark.django_db
def test_start_ghi_started_at_va_submit_tu_tinh_thoi_luong(em, de):
    """Đỏ trên mã cũ: `started_at` có cột nhưng không dòng mã nào ghi vào."""
    mo = _bat_dau(em, de)
    assert mo.status_code == 200
    aid = mo.data['attemptId']
    assert q1("SELECT started_at FROM mock_attempts WHERE id=%s", (aid,))['started_at'] is not None
    assert mo.data['secondsLeft'] == 20 * 60

    # Trình duyệt khai một con số vô lý; máy chủ phải bỏ qua nó. Lùi đồng hồ
    # 5 phút để thời lượng máy chủ tính ra là một con số PHÂN BIỆT ĐƯỢC với 0 —
    # `>= 0` suông là phép kiểm hằng đúng, mọi đường ghi đều đã qua `max(0, …)`.
    _lui_dong_ho(aid, 300)
    r = _nop(em, de, {'q1': '4'}, duration_seconds=-5000)
    assert 300 <= (r.data.get('durationSeconds') or 0) < 360, r.data.get('durationSeconds')
    dong = q1("SELECT duration_seconds, submitted_at FROM mock_attempts WHERE id=%s", (aid,))
    assert 300 <= dong['duration_seconds'] < 360
    assert dong['submitted_at'] is not None


@pytest.mark.django_db
def test_tai_lai_trang_KHONG_duoc_cap_them_gio(em, de):
    """F5 giữa chừng phải NỐI TIẾP lượt đang mở, không cấp lại từ đầu.

    Lùi đồng hồ 5 phút giữa hai lần gọi: một bản mã luôn cấp lại đủ 1200 giây
    sẽ ĐỎ ở đây, còn phép so `b <= a` suông thì nó vẫn lọt (1200 <= 1200).
    """
    a = _bat_dau(em, de)
    _lui_dong_ho(a.data['attemptId'], 300)
    b = _bat_dau(em, de)
    assert b.data['attemptId'] == a.data['attemptId']
    assert 890 <= b.data['secondsLeft'] <= 900, b.data['secondsLeft']


@pytest.mark.django_db
def test_luu_tam_cau_tra_loi_va_lay_lai_duoc_khi_vao_lai(em, de):
    """Lỡ F5 ở phút thứ 15 mà mất trắng bài làm là một hồi quy lên học viên
    thật — đồng hồ vẫn chạy tiếp mà câu trả lời thì về rỗng."""
    _bat_dau(em, de)
    luu = _luu(em, de, {'q1': '4', 'q3': 'thấp'})
    assert luu.data.get('ok') is True
    lai = _bat_dau(em, de)
    assert lai.data.get('savedAnswers') == {'q1': '4', 'q3': 'thấp'}


@pytest.mark.django_db
def test_luu_tam_khong_phai_cua_sau_de_lay_dap_an(em, de):
    """Kiểm theo Ý ĐỊNH (không rò thông tin chấm) chứ không theo danh sách khoá
    cứng: bản cũ dùng `set(r.data) <= {'ok','saved'}` nên nó đỏ ngay khi thêm
    `secondsLeft` — một trường vô hại mà client vốn đã biết."""
    _bat_dau(em, de)
    r = _luu(em, de, {'q1': '4'})
    cam = {'results', 'score', 'total', 'answer', 'answers', 'correct',
           'section_scores', 'weakest'}
    assert not (set(r.data) & cam), r.data


@pytest.mark.django_db
def test_het_gio_thi_van_cham_nhung_KHONG_vao_so(em, de):
    """Mở đề rồi ba ngày sau mới nộp: chấm để xem lại, nhưng không tính điểm."""
    mo = _bat_dau(em, de)
    _lui_dong_ho(mo.data['attemptId'], 3 * 24 * 3600)
    r = _nop(em, de, DUNG_HET)
    assert r.data['score'] == 3
    assert r.data.get('counted') is False
    assert r.data.get('notCountedReason') == 'het_gio'
    assert r.data.get('xpGained') == 0


# ── Đề chưa xuất bản ────────────────────────────────────────────────────────

@pytest.mark.django_db
def test_de_chua_xuat_ban_thi_khong_nop_duoc(em, de_an):
    """Đỏ trên mã cũ: `MockSubmitView` không lọc `is_published`, nên đề đang
    soạn dở vẫn nộp được — và vẫn cộng XP."""
    r = _nop(em, de_an, {'q1': '4'})
    assert r.status_code == 404


@pytest.mark.django_db
def test_de_chua_xuat_ban_thi_khong_bat_dau_duoc(em, de_an):
    r = _bat_dau(em, de_an)
    assert r.status_code == 404


@pytest.mark.django_db
def test_luot_dang_mo_khong_hien_trong_lich_su(em, de):
    """Dòng `submitted_at IS NULL` chưa phải một kết quả."""
    from mockexam.views import MockAttemptsView
    _bat_dau(em, de)
    r = _goi(MockAttemptsView, 'get', ai=em)
    assert [a for a in r.data['attempts'] if a['exam_id'] == de] == []


@pytest.mark.django_db
def test_diem_de_gan_nhat_khong_bi_luot_dang_mo_de_len(em, de):
    """Postgres xếp NULL LÊN ĐẦU trong `ORDER BY ... DESC`: không lọc thì vừa
    bấm "Bắt đầu" là Trang của tôi báo điểm đề gần nhất 0/0."""
    from stats.views import HsaSummaryView
    _bat_dau(em, de)
    _nop(em, de, DUNG_HET)
    _bat_dau(em, de)
    r = _goi(HsaSummaryView, 'get', ai=em)
    assert r.data['lastMockScore'] == 3 and r.data['lastMockTotal'] == 3


# ── Audit chéo 01/09/2026 · nộp muộn KHÔNG được trả lại lượt tính điểm ─────

@pytest.mark.django_db
def test_nop_muon_KHONG_tra_lai_luot_tinh_diem(em, de):
    """Lỗ nặng nhất mà đợt audit bắt được, và là lỗi của chính bản vá L4.

    `/submit` ghi `counted = tinh_diem`, mà `tinh_diem` là FALSE khi nộp muộn →
    cờ bị HẠ → dòng rơi khỏi `uq_mock_attempt_tinh_diem` → lượt tính điểm được
    TRẢ LẠI. Cộng với việc phản hồi trả đáp án cho mọi câu đã điền:

        /start → chờ quá giờ → /submit rác (nhận trọn đáp án VÀ lấy lại lượt)
              → /start → /submit đúng → 9/9 + 100 XP

    Đường `/start` (`_dong_luot_qua_gio`) đã cố ý không đụng cột này ngay từ
    đầu. Hai đường, hai luật, và đường lỏng hơn là đường học viên gọi được.
    """
    a = _bat_dau(em, de)
    assert a.data.get('counts') is True
    _lui_dong_ho(a.data['attemptId'], 20 * 60 + 300)

    tre = _nop(em, de, {'q1': 'bua', 'q2': 'bua', 'q3': 'bua'})
    assert tre.data.get('notCountedReason') == 'het_gio'

    con = q1("SELECT counted FROM mock_attempts WHERE id=%s", (a.data['attemptId'],))
    assert con['counted'] is True, 'nộp muộn đã HẠ cờ và trả lại lượt tính điểm'

    b = _bat_dau(em, de)
    assert b.data.get('counts') is False, 'lượt tính điểm đã bị tiêu, không được cấp lại'
    r = _nop(em, de, DUNG_HET)
    assert r.data.get('counted') is False
    assert r.data.get('xpGained') == 0


@pytest.mark.django_db
def test_luu_tam_SAU_CHUONG_bi_tu_choi(em, de):
    """Không có hàng rào giờ ở `/save` thì nó thành đường vòng TỐT HƠN đường
    trung thực: hết giờ → tra cứu vài ngày → `/save` bộ hoàn hảo → `/start` →
    `_dong_luot_qua_gio` chấm chính bộ ấy và GIỮ NGUYÊN `counted`, tức 9/9 vào
    sổ. Trong khi nộp muộn tử tế qua `/submit` thì không được tính điểm.
    """
    a = _bat_dau(em, de)
    assert _luu(em, de, {'q1': '4'}).status_code == 200
    _lui_dong_ho(a.data['attemptId'], 20 * 60 + 300)

    tre = _luu(em, de, DUNG_HET)
    assert tre.status_code == 409, tre.data
    dong = q1("SELECT answers_json FROM mock_attempts WHERE id=%s", (a.data['attemptId'],))
    import json as _json
    da_luu = dong['answers_json']
    if isinstance(da_luu, str):
        da_luu = _json.loads(da_luu)
    assert da_luu == {'q1': '4'}, 'bài viết sau chuông vẫn ghi được: %s' % (da_luu,)


# ── Nhập đề thi thử từ bảng tính (04/09/2026) ───────────────────────────────
#
# Đo trước khi làm: đúng MỘT đề tồn tại trên CSDL, đến từ `seed_data`, và không
# có API quản trị nào — tức đường duy nhất để có đề thứ hai là sửa mã nguồn.
# Cách làm lấy từ Kahoot: tải mẫu → điền → tải lên → báo lỗi TỪNG DÒNG.

import io as _io  # noqa: E402

import pytest as _pytest  # noqa: E402

_TIEU_DE = ['Phần thi', 'Câu hỏi', 'Lựa chọn A', 'Lựa chọn B', 'Lựa chọn C',
            'Lựa chọn D', 'Đáp án', 'Mã câu', 'Chủ đề', 'Giải thích']


def _xlsx(hang):
    """Dựng một tệp .xlsx trong bộ nhớ từ danh sách dòng."""
    import openpyxl
    wb = openpyxl.Workbook()
    for h in hang:
        wb.active.append(h)
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ban_ghi(hang):
    from common.bangtinh import thanh_ban_ghi
    from mockexam.nhap import COT_BAT_BUOC, TEN_KHAC
    return thanh_ban_ghi(hang, cot_bat_buoc=COT_BAT_BUOC, ten_khac=TEN_KHAC)


def test_dap_an_nguyen_van_THANG_chu_cai():
    """Luật tinh tế nhất của cả đường nhập, và nó có thật trong đề HSA.

    Đáp án nhận hai cách viết: chép nguyên văn phương án đúng, hoặc ghi A/B/C/D.
    Nếu để chữ cái thắng trước thì một câu hỏi VỀ trắc nghiệm — phương án đúng
    là chuỗi "B" — sẽ bị hiểu thành "phương án thứ hai". Sai âm thầm, và chỉ lộ
    ra khi học viên đã thi xong.
    """
    from mockexam.nhap import doc_cau_hoi

    # PHƯƠNG ÁN PHẢI XÁO THỨ TỰ. Bản đầu của phép kiểm này dùng
    # options = ['A','B','C','D'] với đáp án 'B' — một ca SUY BIẾN: đọc "B" là
    # nguyên văn hay là "phương án thứ hai" đều ra cùng chữ B. Đảo thứ tự hai
    # luật trong `nhap.py` rồi chạy lại → VẪN XANH, tức nó không kiểm gì.
    # Xáo đi thì hai cách hiểu tách hẳn: nguyên văn → 'B' (ở vị trí 1),
    # chữ cái → 'A' (phương án thứ hai).
    cau, loi = doc_cau_hoi(_ban_ghi([
        _TIEU_DE,
        ['Định lượng', 'Đáp án đúng của câu 7 là chữ nào?',
         'B', 'A', 'C', 'D', 'B', '', '', ''],
    ]))
    assert not loi, loi
    assert cau[0]['answer'] == 'B', \
        'đọc "B" thành "phương án thứ hai" (%r) — nguyên văn phải thắng' % cau[0]['answer']
    assert cau[0]['options'] == ['B', 'A', 'C', 'D']

    # Còn khi KHÔNG có phương án nào trùng chữ cái thì chữ cái mới được dùng.
    cau2, loi2 = doc_cau_hoi(_ban_ghi([
        _TIEU_DE,
        ['Định lượng', '2+2?', '3', '4', '5', '6', 'B', '', '', ''],
    ]))
    assert not loi2, loi2
    assert cau2[0]['answer'] == '4'


def test_so_dong_bao_loi_khop_voi_excel():
    """Báo "dòng 4" thì người soạn nhấn Ctrl+G là tới nơi; báo "phần tử thứ 3
    của mảng" thì họ phải tự đếm, và đếm sai."""
    from mockexam.nhap import doc_cau_hoi
    _, loi = doc_cau_hoi(_ban_ghi([
        _TIEU_DE,
        ['Định lượng', 'ok', 'a', 'b', '', '', 'a', '', '', ''],   # dòng 2 — đúng
        ['Định lượng', '', 'a', 'b', '', '', 'a', '', '', ''],      # dòng 3 — thiếu câu hỏi
        ['Sai phần', 'x', 'a', 'b', '', '', 'a', '', '', ''],       # dòng 4 — sai phần thi
    ]))
    assert len(loi) == 2, loi
    assert loi[0].startswith('Dòng 3:'), loi[0]
    assert loi[1].startswith('Dòng 4:'), loi[1]


def test_bo_trong_het_lua_chon_thi_thanh_cau_dien():
    from mockexam.nhap import doc_cau_hoi
    cau, loi = doc_cau_hoi(_ban_ghi([
        _TIEU_DE,
        ['Định lượng', 'Điền số còn thiếu: 2,4,8,…', '', '', '', '', '16', '', '', ''],
    ]))
    assert not loi, loi
    assert cau[0]['type'] == 'fill' and 'options' not in cau[0]


def test_hai_phuong_an_trung_nhau_bi_chan():
    """Máy chấm so theo NỘI DUNG phương án, nên hai phương án giống hệt nhau là
    một câu không chấm được — chặn ở đường nhập, đừng để lộ ra lúc thi."""
    from mockexam.nhap import doc_cau_hoi
    _, loi = doc_cau_hoi(_ban_ghi([
        _TIEU_DE, ['Định lượng', 'x', '10', '10', '20', '', '10', '', '', ''],
    ]))
    assert loi and 'trùng nhau' in loi[0], loi


def test_thieu_cot_bat_buoc_noi_ro_thieu_cot_nao():
    from common.bangtinh import LoiBangTinh
    with _pytest.raises(LoiBangTinh) as e:
        _ban_ghi([['Phần thi', 'Câu hỏi'], ['Định lượng', 'x']])
    assert 'đáp án' in str(e.value), str(e.value)


def test_nhap_de_qua_HTTP_tao_de_CHUA_xuat_ban(admin_api, db):
    """Đề nhập vào mặc định CHƯA xuất bản: nhập nhầm cột đáp án mà học viên thấy
    ngay là chấm sai cho cả lớp trước khi ai kịp nhận ra."""
    import json as _json

    from django.core.files.uploadedfile import SimpleUploadedFile

    from common.db import q1

    tep = SimpleUploadedFile('de.xlsx', _xlsx([
        _TIEU_DE,
        ['Định lượng', '2+2?', '3', '4', '5', '6', 'B', '', 'Số học', 'Cộng hai số.'],
        ['Định tính', 'Điền từ còn thiếu', '', '', '', '', 'nhà', '', '', ''],
    ]), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    r = admin_api.post('/api/admin/mock-exams/import',
                       {'file': tep, 'title': 'Đề thử nhập', 'duration_minutes': '30'},
                       format='multipart')
    assert r.status_code == 201, r.json()
    assert r.json()['questions'] == 2

    row = q1('SELECT title, total_questions, is_published, questions_json '
             'FROM mock_exams WHERE id=%s', (r.json()['id'],))
    assert row['is_published'] is False, 'đề mới KHÔNG được tự xuất bản'
    assert row['total_questions'] == 2
    qs = row['questions_json']
    if isinstance(qs, str):
        qs = _json.loads(qs)
    assert qs[0]['answer'] == '4' and qs[0]['section'] == 'quantitative'
    assert qs[1]['type'] == 'fill' and qs[1]['section'] == 'verbal'


def test_khong_ghi_de_len_de_da_co_nguoi_lam(admin_api, db):
    """Đổi đề đã có lượt chấm là làm điểm cũ mất ý nghĩa — chúng chấm trên bộ
    câu hỏi khác. Chặn thẳng và nói rõ đường đi tiếp."""
    from django.core.files.uploadedfile import SimpleUploadedFile

    from common.db import q1

    co_luot = q1('SELECT exam_id FROM mock_attempts WHERE exam_id IS NOT NULL LIMIT 1')
    if not co_luot:
        _pytest.skip('CSDL chưa có lượt thi nào để dựng cảnh')

    tep = SimpleUploadedFile('de.xlsx', _xlsx([
        _TIEU_DE, ['Định lượng', 'x', 'a', 'b', '', '', 'a', '', '', ''],
    ]), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r = admin_api.post('/api/admin/mock-exams/import',
                       {'file': tep, 'title': 'Ghi đè thử', 'duration_minutes': '30',
                        'exam_id': str(co_luot['exam_id'])}, format='multipart')
    assert r.status_code == 409, r.status_code
    assert 'lượt làm bài' in r.json()['error']


def test_hoc_vien_khong_vao_duoc_khu_soan_de(auth_api):
    assert auth_api.get('/api/admin/mock-exams').status_code == 403
    assert auth_api.get('/api/admin/mock-exams/template.xlsx').status_code == 403


def test_mau_tai_ve_doc_nguoc_lai_duoc(admin_api):
    """Mẫu và bộ đọc phải khớp nhau. Sinh mẫu tại chỗ rồi ĐỌC LẠI chính nó — ai
    đổi tên cột ở một bên thì phép kiểm này đỏ ngay, thay vì để TopHSA điền cả
    một ngân hàng câu hỏi rồi mới phát hiện."""
    from common.bangtinh import doc
    from mockexam.nhap import doc_cau_hoi

    r = admin_api.get('/api/admin/mock-exams/template.xlsx')
    assert r.status_code == 200
    hang = doc('mau.xlsx', r.content)
    cau, loi = doc_cau_hoi(_ban_ghi(hang))
    assert not loi, loi
    assert len(cau) == 3, cau


# ── BỐN LỖI CỦA CHÍNH ĐƯỜNG NHẬP NÀY (vá 04/09/2026, một ngày sau khi viết) ──
#
# Cả bốn đều thuộc loại KHÔNG BÁO GÌ. Đường nhập được dựng với đúng một lời hứa
# — "kiểm hết rồi mới ghi, sai thì báo đúng số dòng" — và lỗi im lặng là thứ
# duy nhất phá được lời hứa ấy mà không ai thấy.


def _xlsx_dinh_dang(hang, dinh_dang=None, ten_trang=None):
    """Như `_xlsx` nhưng đặt được `number_format` cho từng ô: {(dòng, cột): fmt}."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    if ten_trang:
        ws.title = ten_trang
    for h in hang:
        ws.append(h)
    for (d, c), fmt in (dinh_dang or {}).items():
        ws.cell(row=d, column=c).number_format = fmt
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_cot_lua_chon_trong_khong_lam_lech_dap_an():
    """① Một cột "Lựa chọn" bỏ trống ở GIỮA làm đáp án chữ cái trỏ sai phương án.

    A, B, (C trống), D — bản đầu nén danh sách còn [A, B, D] rồi vẫn tra
    `ord(chữ)-65` trên nó. Đáp án "C" thành `[A,B,D][2]` = D: KHÔNG báo gì, ghi
    thẳng vào ngân hàng đề, và lộ ra dưới dạng "học viên chọn đúng mà bị trừ
    điểm" nhiều tuần sau, không dấu vết.
    """
    from mockexam.nhap import doc_cau_hoi

    # Đáp án "D" — cột D CÓ nội dung, phải lấy ĐÚNG nội dung cột D.
    cau, loi = doc_cau_hoi(_ban_ghi([
        _TIEU_DE,
        ['Định lượng', 'Chọn đi', 'quả táo', 'quả lê', '', 'quả nho',
         'D', '', '', ''],
    ]))
    assert not loi, loi
    assert cau[0]['answer'] == 'quả nho', cau[0]['answer']
    assert cau[0]['options'] == ['quả táo', 'quả lê', 'quả nho']

    # Đáp án "C" — cột C TRỐNG. Phải BÁO, tuyệt đối không được lặng lẽ lấy D.
    cau2, loi2 = doc_cau_hoi(_ban_ghi([
        _TIEU_DE,
        ['Định lượng', 'Chọn đi', 'quả táo', 'quả lê', '', 'quả nho',
         'C', '', '', ''],
    ]))
    assert loi2, 'đáp án C trỏ vào cột trống mà không báo gì'
    assert 'Lựa chọn C' in loi2[0], loi2
    assert not cau2, 'đã báo lỗi thì không được ghi câu nào'


def test_o_phan_tram_giu_dung_cai_nguoi_soan_nhin_thay():
    """② Ô hiện "30%" — Excel lưu 0.3 — bản đầu trả chuỗi "0.3".

    Không chỗ nào báo lỗi vì "0.3" là một chuỗi hợp lệ; nó chỉ SAI. Người soạn
    gõ 30% vào ô đáp án và học viên phải trả lời "0.3" mới được tính đúng.
    """
    from common.bangtinh import doc

    du = _xlsx_dinh_dang(
        [_TIEU_DE,
         ['Định lượng', 'Lãi suất là bao nhiêu?', '', '', '', '', 0.3, '', '', '']],
        dinh_dang={(2, 7): '0%'})
    hang = doc('de.xlsx', du)
    cot_dap_an = hang[1][6]
    assert cot_dap_an == '30%', 'ô hiện 30%% mà đọc ra %r' % cot_dap_an

    from mockexam.nhap import doc_cau_hoi
    cau, loi = doc_cau_hoi(_ban_ghi(hang))
    assert not loi, loi
    assert cau[0]['answer'] == '30%', cau[0]['answer']


def test_dong_trong_khong_lam_lech_so_dong_bao_loi():
    """③ Số dòng trong thông báo lỗi phải là số dòng NHƯ TRONG EXCEL.

    Bản đầu lọc dòng trống RỒI mới đánh số, nên mỗi dòng trống đẩy toàn bộ phần
    dưới lệch một. Docstring của `thanh_ban_ghi` hứa đúng điều nó không làm — và
    hậu quả không phải là một con số xấu, mà là người soạn mở Excel, nhấn Ctrl+G
    tới dòng được báo, và thấy một dòng KHÔNG có lỗi gì.
    """
    from mockexam.nhap import doc_cau_hoi

    _, loi = doc_cau_hoi(_ban_ghi([
        _TIEU_DE,                                                    # dòng 1
        ['Định lượng', '2+2?', '3', '4', '', '', '4', '', '', ''],   # dòng 2 — đúng
        ['', '', '', '', '', '', '', '', '', ''],                    # dòng 3 — TRỐNG
        ['Định lượng', '', '', '', '', '', '5', '', '', ''],         # dòng 4 — thiếu câu hỏi
    ]))
    assert loi, 'dòng 4 thiếu "câu hỏi" mà không báo'
    assert loi[0].startswith('Dòng 4:'), loi[0]


def test_du_lieu_o_trang_thu_hai_van_doc_duoc():
    """④ Trang 1 là "Hướng dẫn", dữ liệu ở trang 2 — chính hình dạng MẪU của mình.

    `AdminMockExamTemplateView` sinh mẫu có hai trang. Người dùng đảo thứ tự
    hoặc thêm một trang ghi chú lên đầu là chuyện bình thường, và bản đầu cứng
    `worksheets[0]` nên báo "thiếu cột bắt buộc" trong khi dữ liệu nằm ngay
    trang bên cạnh — một thông báo đúng chữ, chỉ tới chỗ không có gì sai.
    """
    import openpyxl

    from common.bangtinh import doc

    wb = openpyxl.Workbook()
    wb.active.title = 'Hướng dẫn'
    wb.active.append(['Điền vào trang "Câu hỏi" nhé.'])
    ws = wb.create_sheet('Câu hỏi')
    ws.append(_TIEU_DE)
    ws.append(['Định lượng', '2+2?', '3', '4', '', '', 'B', '', '', ''])
    buf = _io.BytesIO()
    wb.save(buf)

    from mockexam.nhap import doc_cau_hoi
    cau, loi = doc_cau_hoi(_ban_ghi(doc('de.xlsx', buf.getvalue())))
    assert not loi, loi
    assert cau[0]['answer'] == '4'


def test_exam_id_khong_phai_so_tra_400_chu_khong_no(admin_api, db):
    """⑤ `mock_exams.id` là INTEGER; `exam_id` từ biểu mẫu là chuỗi.

    Chuyền thẳng `'abc'` vào truy vấn thì psycopg ném — tức 500, tức trang báo
    "lỗi máy chủ" cho một việc người dùng gõ sai.
    """
    from django.core.files.uploadedfile import SimpleUploadedFile

    tep = SimpleUploadedFile('de.xlsx', _xlsx([
        _TIEU_DE, ['Định lượng', 'x', 'a', 'b', '', '', 'a', '', '', ''],
    ]), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r = admin_api.post('/api/admin/mock-exams/import',
                       {'file': tep, 'title': 'Đề thử', 'duration_minutes': '30',
                        'exam_id': 'abc'}, format='multipart')
    assert r.status_code == 400, r.status_code
    assert 'exam_id' in r.json()['error']


def test_o_qua_dai_thi_BAO_chu_khong_cat_ngam():
    """⑥ `MAX_O` được khai kèm chú thích "Vượt thì BÁO, không cắt" — và trước
    bản vá này KHÔNG DÒNG NÀO trong repo dùng tới nó. Một lời hứa không có mã
    đứng sau còn tệ hơn không hứa gì."""
    from common.bangtinh import MAX_O, LoiBangTinh

    with _pytest.raises(LoiBangTinh) as e:
        _ban_ghi([_TIEU_DE,
                  ['Định lượng', 'x' * (MAX_O + 1), '', '', '', '', 'a', '', '', '']])
    assert 'Dòng 2' in str(e.value) and 'câu hỏi' in str(e.value), str(e.value)
